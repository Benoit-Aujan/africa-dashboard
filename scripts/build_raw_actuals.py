"""
build_raw_actuals.py
Enriches data.json daily records with raw room counts and revenue from:
  - VPEM_ABAZ_ASLV Pick Up Tracker files (2025 and 2026)
  - Radisson Blu Maputo Revenue Report XLS files (all months)

Adds per-property fields: rooms_occ, rooms_avail, rev_usd
Recomputes occ and adr from those raw figures.
Budget, LE, and PY fields are left untouched.
"""

import json, glob, re, datetime, os
import openpyxl
import xlrd
try:
    import pdfplumber
    _PDF_OK = True
except ImportError:
    _PDF_OK = False
try:
    from pyxlsb import open_workbook as open_xlsb
    _XLSB_OK = True
except ImportError:
    _XLSB_OK = False

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_JSON = os.path.join(BASE, "data", "data.json")
TEMP = "C:/Users/benoit.haas/AppData/Local/Temp"
FX = 65.0  # MZN → USD

# All tracker files in Temp (date-prefixed saves from email + legacy undated copies),
# sorted oldest→newest so newer files win on overlap.
TRACKER_FILES = sorted(
    glob.glob(f"{TEMP}/*VPEM_ABAZ_ASLV*Pick Up Tracker*.xlsx"),
    key=os.path.getmtime,
)

# Exclude the partial March 2026 file; include all others
RADISSON_EXCLUDE = {"Revenue Report 30 March 2026..xls"}

MONTH_MAP = {
    "JAN": 1, "FEB": 2, "MAR": 3, "MARCH": 3,
    "APR": 4, "MAY": 5, "JUN": 6, "JUNE": 6,
    "JUL": 7, "JULY": 7, "AUG": 8, "SEP": 9, "SEPT": 9,
    "OCT": 10, "NOV": 11, "DEC": 12,
}

# prop_key in sheet name → dashboard property key
PROP_NAME = {"VPEM": "Pemba", "ABAZ": "ABAZ", "ASLV": "ASLV"}


# ── VPEM / ABAZ / ASLV trackers ──────────────────────────────────────────────

def parse_tracker(path):
    """
    Returns {prop: {date_str: {rooms_occ, rooms_avail, rev_usd}}}
    VPEM  → Total Property cols: occ=9, occ%=10, rev=12  (MZN)
    ABAZ  → cols: occ=1, occ%=2, rev=4                   (MZN)
    ASLV  → cols: occ=1, occ%=2, rev=4                   (USD)
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}

    for shname in wb.sheetnames:
        m = re.match(r"^(VPEM|ABAZ|ASLV)\s+(\w+?)(\d{2})$", shname)
        if not m:
            continue
        prop_key, month_str, year_str = m.group(1), m.group(2).upper(), m.group(3)
        month_num = MONTH_MAP.get(month_str)
        if not month_num:
            continue

        year = 2000 + int(year_str)
        prop = PROP_NAME[prop_key]
        is_mzn = (prop_key != "ASLV")

        sh = wb[shname]
        all_rows = list(sh.iter_rows(min_row=3, values_only=True))

        # First pass: infer property capacity from non-zero occupancy rows.
        # Needed so zero-occ days (genuine closures) still contribute to the
        # rooms_avail denominator in weighted Occ calculations.
        occ_col = 9 if prop_key == "VPEM" else 1
        pct_col = 10 if prop_key == "VPEM" else 2
        capacity = 0
        for row in all_rows:
            if not row or row[0] is None: continue
            occ_v = row[occ_col]; pct_v = row[pct_col]
            if occ_v and pct_v and pct_v > 0:
                capacity = max(capacity, round(int(occ_v) / pct_v))

        # Second pass: build daily records
        for row in all_rows:
            if not row or row[0] is None:
                continue
            date = row[0]
            if not isinstance(date, datetime.datetime):
                continue
            # Year-rollover correction (e.g. ABAZ FEB26 with 2025 dates)
            if date.year != year:
                try:
                    date = date.replace(year=year)
                except ValueError:
                    continue
            date_str = date.strftime("%Y-%m-%d")

            if prop_key == "VPEM":
                rooms_occ = row[9]
                occ_pct   = row[10]
                rev       = row[12]
            else:  # ABAZ or ASLV
                rooms_occ = row[1]
                occ_pct   = row[2]
                rev       = row[4]

            if rooms_occ is None or occ_pct is None or rev is None:
                continue

            # Zero-occupancy day: only store for ASLV (boutique lodge with genuine closures).
            # VPEM and ABAZ are city/resort hotels — zero-occ rows are future empty tracker rows.
            if rooms_occ == 0 and rev == 0:
                if prop_key == "ASLV" and capacity > 0:
                    result.setdefault(prop, {})[date_str] = {
                        "rooms_occ": 0, "rooms_avail": capacity, "rev_usd": 0.0,
                    }
                continue

            try:
                rooms_occ   = int(rooms_occ)
                rooms_avail = round(rooms_occ / occ_pct) if occ_pct > 0 else capacity
                rev_usd     = float(rev) / FX if is_mzn else float(rev)
            except (TypeError, ZeroDivisionError):
                continue

            if rooms_avail == 0:
                continue

            result.setdefault(prop, {})[date_str] = {
                "rooms_occ":   rooms_occ,
                "rooms_avail": rooms_avail,
                "rev_usd":     round(rev_usd, 4),
            }

    return result


# ── Radisson Revenue Report XLS ───────────────────────────────────────────────

def _find_rows(sh):
    """Search col B for room-stat labels; return {normalised_label: row_index}."""
    targets = {
        "COMP ROOMS", "MAINT ROOMS", "VACANT ROOMS",
        "ROOMS SOLD", "HOUSE USE ROOMS",
    }
    found = {}
    for r in range(sh.nrows):
        raw = sh.cell_value(r, 1)
        if not isinstance(raw, str):
            continue
        norm = " ".join(raw.strip().upper().split())  # collapse double spaces
        if norm in targets:
            found[norm] = r
    return found


def _num(val, default=0.0):
    """Safely cast an xlrd cell value to float."""
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def parse_radisson(path):
    """Returns {date_str: {rooms_occ, rooms_avail, rev_usd}}"""
    wb  = xlrd.open_workbook(path)
    sh  = wb.sheet_by_name("Daily Input")
    rows = _find_rows(sh)
    REV_ROW = 37  # row 38, 0-indexed

    result = {}
    for c in range(2, sh.ncols):
        date_val = sh.cell_value(0, c)
        if not isinstance(date_val, float) or date_val == 0:
            continue
        try:
            date = datetime.date(*xlrd.xldate_as_tuple(date_val, wb.datemode)[:3])
            date_str = date.strftime("%Y-%m-%d")
        except Exception:
            continue

        def _v(label):
            r = rows.get(label)
            return _num(sh.cell_value(r, c)) if r is not None else 0.0

        comp   = _v("COMP ROOMS")
        maint  = _v("MAINT ROOMS")
        vacant = _v("VACANT ROOMS")
        sold   = _v("ROOMS SOLD")
        house  = _v("HOUSE USE ROOMS")
        rev    = _num(sh.cell_value(REV_ROW, c))

        rooms_occ   = int(round(sold + comp + house))
        rooms_avail = int(round(comp + maint + vacant + sold + house))

        if rooms_avail == 0 or rev == 0:
            continue

        result[date_str] = {
            "rooms_occ":   rooms_occ,
            "rooms_avail": rooms_avail,
            "rev_usd":     round(rev / FX, 4),
        }
    return result


# ── Radisson PDF Flash Report ─────────────────────────────────────────────────

def parse_radisson_pdf(path):
    """
    Parses a Radisson Blu Maputo Daily Manager Flash Report PDF.
    Returns {date_str: {rooms_occ, rooms_avail, rev_usd}} or {}.

    PDF columns (left to right): DAY_2026, MTD_2026, YTD_2026, DAY_2025, MTD_2025, YTD_2025
    We use the DAY_2026 column (first numeric value on each row).

    Date is taken from the "Filter Calendar/Month to Date DD.MM.YY" footer,
    which gives the last date included in the MTD column — the same date as DAY.

    Revenue is in MZN; converted to USD using the FX constant.
    """
    if not _PDF_OK:
        print("  WARNING: pdfplumber not installed — pip install pdfplumber")
        return {}

    text = ""
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
    except Exception as e:
        print(f"  ERROR reading PDF {os.path.basename(path)}: {e}")
        return {}

    # Date from "Filter Calendar/Month to Date DD.MM.YY"
    m = re.search(r"Filter Calendar/Month to Date\s+(\d{2})\.(\d{2})\.(\d{2})", text)
    if not m:
        print(f"  WARNING: no date found in {os.path.basename(path)}")
        return {}
    day, month, yr2 = int(m.group(1)), int(m.group(2)), int(m.group(3))
    date_str = f"{2000 + yr2}-{month:02d}-{day:02d}"

    def _day_val(label):
        """Return the first numeric token after `label` at line start (DAY_2026 column)."""
        hit = re.search(
            rf"^{re.escape(label)}\s+([\d,]+(?:\.\d+)?)",
            text, re.MULTILINE,
        )
        return float(hit.group(1).replace(",", "")) if hit else None

    total_rooms = _day_val("Total Rooms in Hotel")
    rooms_occ   = _day_val("Rooms Occupied")
    room_rev    = _day_val("Room Revenue")

    if None in (total_rooms, rooms_occ, room_rev):
        print(f"  WARNING: missing fields in {os.path.basename(path)} "
              f"(rooms={rooms_occ}, avail={total_rooms}, rev={room_rev})")
        return {}

    return {
        date_str: {
            "rooms_occ":   int(round(rooms_occ)),
            "rooms_avail": int(round(total_rooms)),
            "rev_usd":     round(room_rev / FX, 4),
        }
    }


# ── ABAZ Daily Income .xlsb ───────────────────────────────────────────────────

def parse_abaz_daily(path):
    """
    Parses an ABAZ-Daily Income-YYYY.MM.DD.xlsb file.
    Returns {date_str: {rooms_occ, rooms_avail, rev_usd}} or {}.

    Sheet "Rooms Drivers":
      - Searches for the "Today" row to get Rooms Inventory (avail) and Rooms Sold.
      - Searches for Complementary and House Use rows (those without a 3-letter
        market-segment code in col 2) to get comp/house room counts.
    Sheet "Daily Income":
      - Finds the "Rooms" revenue row for today's MZN total.
    Currency: always MZN → USD via FX constant.
    """
    if not _XLSB_OK:
        print("  WARNING: pyxlsb not installed — cannot parse .xlsb files. Run: pip install pyxlsb")
        return {}

    # Date from filename: ABAZ-Daily Income-2026.04.21.xlsb (may have _N suffix)
    m = re.search(r"(\d{4})\.(\d{2})\.(\d{2})", path)
    if not m:
        return {}
    date_str = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

    try:
        with open_xlsb(path) as wb:
            # ── Rooms Drivers sheet ──
            rooms_sold = comp = house = rooms_avail = None
            with wb.get_sheet("Rooms Drivers") as sh:
                rows = list(sh.rows())

            for row in rows:
                vals = [c.v for c in row]
                if len(vals) < 5:
                    continue
                label = vals[1]
                code  = vals[2]   # market-segment code like "CMP", "HOU", or None

                if label == "Today" and isinstance(vals[3], (int, float)) and vals[3] > 0:
                    # Row: Today | Rooms Inventory | Rooms Sold | % Occ | ...
                    rooms_avail = vals[3]
                    rooms_sold  = vals[4]

                elif label == "Complementary" and code is None:
                    # The Comp & House section row (no 3-letter code), col 3 = today's comp rooms
                    if isinstance(vals[3], (int, float)):
                        comp = vals[3]

                elif label == "House Use" and code is None:
                    if isinstance(vals[3], (int, float)):
                        house = vals[3]

            # ── Daily Income sheet — Rooms revenue today (MZN) ──
            rev_mzn = None
            with wb.get_sheet("Daily Income") as sh:
                for row in sh.rows():
                    vals = [c.v for c in row]
                    if len(vals) > 3 and vals[1] == "Rooms" and isinstance(vals[3], (int, float)):
                        rev_mzn = vals[3]
                        break

    except Exception as e:
        print(f"  ERROR parsing {os.path.basename(path)}: {e}")
        return {}

    if rooms_avail is None or rooms_sold is None or rev_mzn is None:
        print(f"  WARNING: could not extract all fields from {os.path.basename(path)}")
        return {}

    rooms_occ   = int(round(rooms_sold + (comp or 0) + (house or 0)))
    rooms_avail = int(round(rooms_avail))
    rev_usd     = round(rev_mzn / FX, 4)

    if rooms_avail == 0 or rev_usd == 0:
        return {}

    return {date_str: {"rooms_occ": rooms_occ, "rooms_avail": rooms_avail, "rev_usd": rev_usd}}


# ── VPEM (Pemba) Daily Income .xlsb ──────────────────────────────────────────

# Known physical room inventory for VPEM (Avani Pemba) including Residences.
# Used as rooms_avail when reading individual daily xlsb files.
VPEM_CAPACITY = 168

def parse_vpem_daily(path):
    """
    Parses a VPEM-Daily Income-YYYY.MM.DD.xlsb file sent by Hamilton Pasipamire.
    Returns {date_str: {rooms_occ, rooms_avail, rev_usd}} or {}.

    Sheet "Contents":  Year/Month/Day in col 14 (0-indexed) of rows 11/12/13.
    Sheet "Rooms Drivers":
      - Row with label "Total Rooms + Residences (ex. Comp & House)":
            col 3 = today rooms sold (excl comp/house),  col 4 = today rev MZN
      - Row "Complementary"       (code CMP): col 3 = comp rooms today
      - Row "Complimentary Owner" (code COO): col 3 = comp owner rooms today
      - Row "House Use"           (code HOU): col 3 = house use rooms today
    rooms_occ  = sold + comp + comp_owner + house
    rooms_avail = VPEM_CAPACITY (physical inventory)
    rev_usd     = rev_mzn / FX
    """
    if not _XLSB_OK:
        print("  WARNING: pyxlsb not installed — cannot parse VPEM xlsb")
        return {}

    # Date from filename first (more reliable than sheet content)
    m = re.search(r"(\d{4})\.(\d{2})\.(\d{2})", path)
    if m:
        date_str = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    else:
        # Fall back to Contents sheet
        try:
            with open_xlsb(path) as wb:
                with wb.get_sheet("Contents") as sh:
                    rows = list(sh.rows())
                year  = int(rows[11][14].v) if len(rows) > 11 else None
                month = int(rows[12][14].v) if len(rows) > 12 else None
                day   = int(rows[13][14].v) if len(rows) > 13 else None
                if not all([year, month, day]):
                    return {}
                date_str = f"{year}-{month:02d}-{day:02d}"
        except Exception:
            return {}

    try:
        with open_xlsb(path) as wb:
            with wb.get_sheet("Rooms Drivers") as sh:
                rows = list(sh.rows())

        sold = comp = comp_owner = house = rev_mzn = None
        for row in rows:
            vals = [c.v for c in row]
            if len(vals) < 5:
                continue
            label = vals[1]
            if label == "Total Rooms + Residences (ex. Comp & House)":
                if isinstance(vals[3], (int, float)):
                    sold    = vals[3]
                    rev_mzn = vals[4] if isinstance(vals[4], (int, float)) else None
            elif label == "Complementary":
                if isinstance(vals[3], (int, float)):
                    comp = vals[3]
            elif label == "Complimentary Owner":
                if isinstance(vals[3], (int, float)):
                    comp_owner = vals[3]
            elif label == "House Use":
                if isinstance(vals[3], (int, float)):
                    house = vals[3]

        if sold is None or rev_mzn is None:
            return {}

        rooms_occ = int(round(sold + (comp or 0) + (comp_owner or 0) + (house or 0)))
        rev_usd   = round(rev_mzn / FX, 4)

        return {
            date_str: {
                "rooms_occ":   rooms_occ,
                "rooms_avail": VPEM_CAPACITY,
                "rev_usd":     rev_usd,
            }
        }
    except Exception as e:
        print(f"  ERROR parsing VPEM xlsb {os.path.basename(path)}: {e}")
        return {}


# ── Management Pack parsers ───────────────────────────────────────────────────
#
# Hotel Finance Template .xlsb (ABAZ / ASLV / VPEM) — identical 74-sheet template.
# Sheet "Monthly" → LE figures for the current reporting month.
# Sheet "Budget"  → Budget figures for all 12 months of the year.
#
# Row indices (0-based) for the data we need:
#   Row 6  = Room Revenue         Row 7  = Residence Revenue
#   Row 94 = Room Available       Row 95 = Residence Available
#   Row 96 = Rooms Sold (w/o Comp) Row 97 = Residence Sold
#   Row 99 = Complimentary Rooms
#
# Column indices (0-based) per calendar month (quarterly totals are skipped):
MGMT_PACK_MONTH_COLS = {
    1: 3,  2: 4,  3: 5,          # Jan Feb Mar
    4: 8,  5: 9,  6: 10,         # Apr May Jun
    7: 13, 8: 14, 9: 15,         # Jul Aug Sep
    10: 18, 11: 19, 12: 20,      # Oct Nov Dec
}


def _safe_xlsb(rows, row_idx, col_idx):
    """Return float from an xlsb row list, or None on any error."""
    try:
        v = rows[row_idx][col_idx].v
        return float(v) if isinstance(v, (int, float)) else None
    except (IndexError, TypeError, AttributeError):
        return None


def parse_mgmt_pack_minor(path, reporting_month, is_mzn=True):
    """
    Parses an ABAZ / ASLV / VPEM Hotel Finance Template .xlsb management pack.

    reporting_month : datetime.date with day=1 for the month being reported.
    is_mzn          : True for MZN-denominated properties (ABAZ, Pemba);
                      False for USD-denominated (ASLV).

    Revenue in the template is in THOUSANDS of local currency.
    Conversion: MZN-thousands → USD  =  value * 1000 / FX
                USD-thousands → USD  =  value * 1000

    Returns:
      {
        "LE":     { "YYYY-MM": {"occ": float, "adr": float} },   # current month only
        "Budget": { "YYYY-MM": {"occ": float, "adr": float} },   # all 12 months
      }

    OCC = (rooms_sold + res_sold + comp) / (room_avail + res_avail)
    ADR = (room_rev + res_rev)           / (rooms_sold + res_sold + comp)
    """
    if not _XLSB_OK:
        print("  WARNING: pyxlsb not installed — cannot parse management pack xlsb")
        return {}

    rev_scale = (1000.0 / FX) if is_mzn else 1000.0

    year      = reporting_month.year
    cur_month = reporting_month.month
    results   = {"LE": {}, "Budget": {}}

    def _month_vals(rows, month_num):
        col = MGMT_PACK_MONTH_COLS.get(month_num)
        if col is None:
            return None, None
        room_rev   = (_safe_xlsb(rows, 6,  col) or 0) * rev_scale
        res_rev    = (_safe_xlsb(rows, 7,  col) or 0) * rev_scale
        room_avail = _safe_xlsb(rows, 94, col) or 0
        res_avail  = _safe_xlsb(rows, 95, col) or 0
        rooms_sold = _safe_xlsb(rows, 96, col) or 0
        res_sold   = _safe_xlsb(rows, 97, col) or 0
        comp       = _safe_xlsb(rows, 99, col) or 0

        total_avail = room_avail + res_avail
        total_occ   = rooms_sold + res_sold + comp
        total_rev   = room_rev + res_rev

        if total_avail <= 0 or total_occ <= 0:
            return None, None
        occ = round(total_occ / total_avail, 6)
        adr = round(total_rev / total_occ,   4) if total_occ > 0 else None
        return occ, adr

    try:
        with open_xlsb(path) as wb:
            # ── Monthly sheet → LE for ALL months ─────────────────────────────
            # Jan–(cur_month-1) show actuals (label "Actual"); cur_month onwards
            # are forecasts ("Forecast"). We store all of them as LE so that
            # future months get preliminary figures from the current pack, which
            # are later overwritten when those months' own packs arrive.
            try:
                with wb.get_sheet("Monthly") as sh:
                    monthly_rows = list(sh.rows())
                for mon in range(1, 13):
                    occ, adr = _month_vals(monthly_rows, mon)
                    if occ is not None:
                        results["LE"][f"{year}-{mon:02d}"] = {"occ": occ, "adr": adr}
            except Exception as e:
                print(f"  WARNING: 'Monthly' sheet error in {os.path.basename(path)}: {e}")

            # ── Budget sheet → all 12 months ──────────────────────────────────
            try:
                with wb.get_sheet("Budget") as sh:
                    budget_rows = list(sh.rows())
                for mon in range(1, 13):
                    occ, adr = _month_vals(budget_rows, mon)
                    if occ is not None:
                        results["Budget"][f"{year}-{mon:02d}"] = {"occ": occ, "adr": adr}
            except Exception as e:
                print(f"  WARNING: 'Budget' sheet error in {os.path.basename(path)}: {e}")

    except Exception as e:
        print(f"  ERROR parsing management pack {os.path.basename(path)}: {e}")
        return {}

    print(f"  Mgmt pack minor parsed: LE={len(results['LE'])} months, "
          f"Budget={len(results['Budget'])} months")
    return results


def parse_mgmt_pack_radisson(path, reporting_month):
    """
    Parses the MPMZH USAH 12MONTHS REPORT .xlsx management pack.

    File: "7.MPMZH USAH 12MONTHS REPORT <MONTH> <YEAR>.xlsx"  (single "Data" sheet)
    Revenue: MZN (Local Currency), in thousands -> converted with * 1000 / FX.

    Layout (173 rows x 149 cols):
      - Column M (0-based index 12): row labels
      - Row 14:  ROOMS REVENUE
      - Row 111: TOTAL ROOMS AVAILABLE
      - Row 113: TOTAL ROOMS OCCUPIED
      - Row 12:  Month labels "YYYY.MON" (e.g. "2026.JAN") -- one per 8-column block
      - 8-column block (offsets relative to month-label column):
            +0  ACTUAL revenue
            +1  BUDGET_4 revenue      <- Budget revenue
            +2  spacer                <- Budget Avail/Occ
            +3  ACTUAL(LY) revenue
            +4  spacer                <- LY Actual Avail/Occ
            +5  FORECAST_12 revenue   <- LE revenue
            +6  spacer                <- LE Avail/Occ
            +7  spacer (empty)

    Returns:
      {
        "LE":     { "YYYY-MM": {"occ": float, "adr": float} },
        "Budget": { "YYYY-MM": {"occ": float, "adr": float} },
      }
    """
    results = {"LE": {}, "Budget": {}}

    def _sf(row, col):
        try:
            v = row[col]
            return float(v) if v is not None else None
        except (IndexError, TypeError, ValueError):
            return None

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Data"]
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()

        if not all_rows:
            return {}

        # ── Find metric rows by scanning column M (0-based index 12) ──────────
        LABEL_COL = 12
        rev_row = avail_row = occ_row = None
        for i, row in enumerate(all_rows):
            if len(row) <= LABEL_COL or row[LABEL_COL] is None:
                continue
            lbl = str(row[LABEL_COL]).strip().upper()
            if "ROOMS REVENUE" in lbl and "TOTAL" not in lbl and rev_row is None:
                rev_row = i
            elif "TOTAL ROOMS AVAILABLE" in lbl and avail_row is None:
                avail_row = i
            elif "TOTAL ROOMS OCCUPIED" in lbl and occ_row is None:
                occ_row = i
            if rev_row is not None and avail_row is not None and occ_row is not None:
                break

        if None in (rev_row, avail_row, occ_row):
            print(f"  WARNING: Missing metric rows: rev={rev_row} avail={avail_row} "
                  f"occ={occ_row} in {os.path.basename(path)}")
            return {}

        # ── Find month labels row ─────────────────────────────────────────────
        # The correct row has "YYYY.MON" cells at regular 8-column intervals,
        # all in the same year (the reporting year).
        # Strategy: pick the row with the most matches WHERE THE MAJORITY ARE
        # in the reporting year (not historical).  Row 12 in this template.
        year = reporting_month.year
        month_labels_row_idx = None
        best_score = 0
        for i, row in enumerate(all_rows[:20]):
            rpt_year_count = sum(
                1 for v in row
                if v and re.match(rf"{year}\.\w{{3,5}}$", str(v).strip())
            )
            if rpt_year_count > best_score:
                best_score = rpt_year_count
                month_labels_row_idx = i
        if month_labels_row_idx is None or best_score < 3:
            print(f"  WARNING: Could not find {year} month labels row in {os.path.basename(path)}")
            return {}

        # Map column -> (year, month_num)
        col_to_month = {}
        for j, cell in enumerate(all_rows[month_labels_row_idx]):
            if cell is None:
                continue
            m = re.match(r"(\d{4})\.(\w{3,5})$", str(cell).strip())
            if m:
                yr, mon = int(m.group(1)), MONTH_MAP.get(m.group(2).upper())
                if mon:
                    col_to_month[j] = (yr, mon)

        if not col_to_month:
            return {}

        # ── Extract Budget and LE for each month block ─────────────────────────
        # Month-label column = block start (ACTUAL).
        # Offsets: +1=BUDGET_4 rev, +2=Budget rooms, +5=LE rev, +6=LE rooms
        for b_start, (yr, mon) in col_to_month.items():
            if not 1 <= mon <= 12:
                continue
            month_str = f"{yr}-{mon:02d}"

            # Budget
            bgt_rev   = _sf(all_rows[rev_row],   b_start + 1)
            bgt_avail = _sf(all_rows[avail_row], b_start + 2)
            bgt_occ   = _sf(all_rows[occ_row],   b_start + 2)
            if bgt_rev and bgt_avail and bgt_occ and bgt_avail > 0 and bgt_occ > 0:
                results["Budget"][month_str] = {
                    "occ": round(bgt_occ / bgt_avail, 6),
                    "adr": round(bgt_rev * 1000 / FX / bgt_occ, 4),
                }

            # LE (FORECAST_12)
            le_rev   = _sf(all_rows[rev_row],   b_start + 5)
            le_avail = _sf(all_rows[avail_row], b_start + 6)
            le_occ   = _sf(all_rows[occ_row],   b_start + 6)
            if le_rev and le_avail and le_occ and le_avail > 0 and le_occ > 0:
                results["LE"][month_str] = {
                    "occ": round(le_occ / le_avail, 6),
                    "adr": round(le_rev * 1000 / FX / le_occ, 4),
                }

    except Exception as e:
        print(f"  ERROR parsing Radisson mgmt pack {os.path.basename(path)}: {e}")
        return {}

    print(f"  Mgmt pack Radisson parsed: LE={len(results['LE'])} months, "
          f"Budget={len(results['Budget'])} months")
    return results


# ── Merge into data.json ──────────────────────────────────────────────────────

def main():
    # 1. Load existing data.json
    with open(DATA_JSON, encoding="utf-8") as f:
        data = json.load(f)

    # 2. Build actuals dict from tracker files
    #    actuals[prop][date_str] = {rooms_occ, rooms_avail, rev_usd}
    actuals = {}
    for path in TRACKER_FILES:
        try:
            tracker = parse_tracker(path)
        except Exception as e:
            print(f"  SKIP tracker {os.path.basename(path)}: {e}")
            continue
        for prop, days in tracker.items():
            prop_store = actuals.setdefault(prop, {})
            prop_store.update(days)  # later file (2026) overwrites if overlap
    print(f"Tracker actuals loaded: "
          + ", ".join(f"{p}={len(d)}" for p, d in actuals.items()))

    # 3. Build Radisson actuals
    rad_actuals = {}
    rad_files = sorted(
        f for f in glob.glob(f"{TEMP}/Revenue Report*.xls")
        if os.path.basename(f) not in RADISSON_EXCLUDE
    )
    for path in rad_files:
        name = os.path.basename(path)
        try:
            days = parse_radisson(path)
            rad_actuals.update(days)
            print(f"  Radisson {name}: {len(days)} days")
        except Exception as e:
            print(f"  SKIP {name}: {e}")
    # Also load PDF flash reports for dates not yet covered by XLS files.
    # XLS Revenue Reports are more detailed, so XLS always takes priority.
    rad_pdf_files = sorted(
        glob.glob(f"{TEMP}/*manager_report*.pdf"),
        key=os.path.getmtime,
    )
    pdf_added = 0
    for path in rad_pdf_files:
        name = os.path.basename(path)
        try:
            days = parse_radisson_pdf(path)
            for date_str, raw in days.items():
                if date_str not in rad_actuals:  # XLS wins on overlap
                    rad_actuals[date_str] = raw
                    pdf_added += 1
                    print(f"  Radisson PDF {name}: added {date_str}")
        except Exception as e:
            print(f"  SKIP PDF {name}: {e}")

    actuals["Radisson"] = rad_actuals
    print(f"Radisson total: {len(rad_actuals)} days ({pdf_added} from PDF)")

    # 4. Update existing records AND insert new dates not yet in data.json
    daily = data.get("daily", {})
    # Capture the original start date BEFORE any insertions
    range_start = min(daily.keys()) if daily else "2025-06-01"

    def _apply_raw(prop_rec, raw):
        ro, ra, ru = raw["rooms_occ"], raw["rooms_avail"], raw["rev_usd"]
        prop_rec["rooms_occ"]   = ro
        prop_rec["rooms_avail"] = ra
        prop_rec["rev_usd"]     = ru
        prop_rec["occ"] = round(ro / ra, 6) if ra > 0 else None
        prop_rec["adr"] = round(ru / ro, 4) if ro > 0 else None

    updated = {p: 0 for p in actuals}
    inserted = 0

    # Update existing records
    for date_str, day_rec in daily.items():
        for prop, days in actuals.items():
            if date_str in days:
                _apply_raw(day_rec.setdefault(prop, {}), days[date_str])
                updated[prop] += 1

    # Insert new dates that have at least one property with raw data,
    # but only within or after the existing date range (don't backfill history
    # before the first Africa weekly report, which had Budget/LE/PY data).
    all_actuals_dates = set()
    for prop_days in actuals.values():
        all_actuals_dates.update(prop_days.keys())

    for date_str in sorted(all_actuals_dates - set(daily.keys())):
        if date_str < range_start:
            continue
        new_rec = {}
        for prop, days in actuals.items():
            if date_str in days:
                prop_rec = {"occ_le": None, "occ_bgt": None, "occ_py": None,
                            "adr_le": None, "adr_bgt": None, "adr_py": None}
                _apply_raw(prop_rec, days[date_str])
                new_rec[prop] = prop_rec
        if new_rec:
            daily[date_str] = new_rec
            inserted += 1

    # Re-sort daily by date
    data["daily"] = dict(sorted(daily.items()))

    print("\nDaily records updated:")
    for p, n in updated.items():
        print(f"  {p}: {n}")
    print(f"New dates inserted: {inserted}")

    # 4.5 Populate PY (prior year) for 2026 dates from 2025 actuals.
    # PY for 2025 dates (= 2024 actuals) does not exist — kept null.
    # First clear any stale PY values (e.g. from old Africa weekly reports),
    # then repopulate 2026 PY from the same tracker/Radisson source files.
    PROPS_PY = ["ABAZ", "Pemba", "ASLV", "Radisson"]
    for date_str, day_rec in daily.items():
        for prop in PROPS_PY:
            rec = day_rec.get(prop)
            if rec is not None:
                rec["occ_py"] = None
                rec["adr_py"] = None

    py_filled = {p: 0 for p in PROPS_PY}
    for date_str, day_rec in daily.items():
        if not date_str.startswith("2026"):
            continue
        py_date = "2025" + date_str[4:]  # e.g. "2026-03-15" -> "2025-03-15"
        for prop in PROPS_PY:
            prop_rec = day_rec.get(prop)
            if prop_rec is None:
                continue
            py_raw = actuals.get(prop, {}).get(py_date)
            if py_raw is None:
                continue
            ro, ra, ru = py_raw["rooms_occ"], py_raw["rooms_avail"], py_raw["rev_usd"]
            prop_rec["occ_py"] = round(ro / ra, 6) if ra > 0 else None
            prop_rec["adr_py"] = round(ru / ro, 4) if ro > 0 else None
            py_filled[prop] += 1

    print("\nPY (2025 actuals) filled for 2026 dates:")
    for p, n in py_filled.items():
        print(f"  {p}: {n}")

    # 5. Normalize benchmarks: LE and Budget are monthly targets — enforce a
    #    single consistent value per property per month across all days.
    #    PY is excluded: it is daily actual data, not a monthly target.
    BENCH_FIELDS = ["occ_le", "occ_bgt", "adr_le", "adr_bgt"]
    PROPS_ALL = ["ABAZ", "Pemba", "ASLV", "Radisson"]

    # Collect all non-null values per (prop, month, field)
    monthly_vals = {}
    for date_str, day_rec in daily.items():
        month = date_str[:7]
        for prop in PROPS_ALL:
            prop_rec = day_rec.get(prop, {})
            for field in BENCH_FIELDS:
                v = prop_rec.get(field)
                if v is not None:
                    monthly_vals.setdefault((prop, month, field), []).append(v)

    # Compute the monthly benchmark from weekly reports (average of available values)
    monthly_bench = {k: sum(vs) / len(vs) for k, vs in monthly_vals.items()}

    # Override with management pack benchmarks stored in data.json["monthly_benchmarks"].
    # Management pack figures are the official monthly close — they take precedence over
    # weekly pick-up report values.
    FIELD_MAP = {"occ_le": "occ", "adr_le": "adr", "occ_bgt": "occ", "adr_bgt": "adr"}
    BENCH_SRC  = {"occ_le": "LE", "adr_le": "LE", "occ_bgt": "Budget", "adr_bgt": "Budget"}
    mgmt_raw = data.get("monthly_benchmarks", {})
    mgmt_override = 0
    for prop in PROPS_ALL:
        prop_mb = mgmt_raw.get(prop, {})
        for month_str, mb in prop_mb.items():
            for bench_field in BENCH_FIELDS:
                src_section = BENCH_SRC[bench_field]
                src_key     = FIELD_MAP[bench_field]
                v = mb.get(src_section, {}).get(src_key) if isinstance(mb.get(src_section), dict) else None
                if v is not None:
                    monthly_bench[(prop, month_str, bench_field)] = v
                    mgmt_override += 1
    if mgmt_override:
        print(f"Management pack overrides applied: {mgmt_override} benchmark cells")

    # Write back to every day
    bench_applied = 0
    for date_str, day_rec in daily.items():
        month = date_str[:7]
        for prop in PROPS_ALL:
            prop_rec = day_rec.get(prop)
            if prop_rec is None:
                continue
            for field in BENCH_FIELDS:
                key = (prop, month, field)
                if key in monthly_bench:
                    prop_rec[field] = round(monthly_bench[key], 6)
                    bench_applied += 1

    print(f"Benchmark cells normalised: {bench_applied}")

    # 6. Save
    with open(DATA_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, separators=(",", ":"))
    print(f"\nSaved -> {DATA_JSON}")


if __name__ == "__main__":
    main()
